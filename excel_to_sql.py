from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import mysql.connector
from mysql.connector import Error
from logger import logger


class ExcelToSQL:
    def __init__(self, filename) -> None:
        self.wb = load_workbook(filename)

    def get_number_of_rows(self, sheet):
        nrows = sheet.max_row
        if nrows > 5000:
            nrows = 5000

        lastrow = 0
        while True:
            if sheet.cell(nrows, 3).value != None:
                lastrow = nrows
                break
            else:
                nrows -= 1
        return lastrow

    def get_sheet(self, sheet_name):
        return self.wb[sheet_name]

    def get_prepared_sheet(self, sheet_name: str, consider_columns: dict):
        sheet = self.get_sheet(sheet_name)
        list_of_rows = []
        num_of_columns = len(consider_columns)
        num_of_rows = self.get_number_of_rows(sheet)
        available_columns = consider_columns.keys()
        unique_columns = {}
        for column in available_columns:
            unique_columns[column] = []
        for row in range(1, num_of_rows):
            column_dict = {}
            for sheet_column, validations in consider_columns.items():
                sheet_column = int(sheet_column)
                column = get_column_letter(sheet_column)

                cellValue = (sheet[column+str(row)].value)
                should_continue = True
                if 'unique' in validations.keys() and validations['unique']:
                    if(cellValue not in unique_columns[str(sheet_column)]):
                        unique_columns[str(sheet_column)].append(cellValue)
                    else:
                        should_continue = False
                if(not should_continue):
                    break
                target_column_name = validations['target_column_name']
                column_dict[target_column_name] = cellValue
            list_of_rows.append(column_dict)

        return list_of_rows

    # @staticmethod
    def set_cursor(self, host, database, user, password):
        self.connection = mysql.connector.connect(
            host=host, database=database, user=user, password=password)
        self.cursor = self.connection.cursor()

    def insert_into_table(self, host, database, user, password, table_name, prepared_list, truncate=False):

        self.set_cursor(host, database, user, password)

        if(truncate):
            query = f"TRUNCATE `{table_name}`;"
            self.cursor.execute(query)
            self.connection.commit()
            msg = f"TRUNCATE ROW AFFECTED = {self.cursor.rowcount}"
            print(msg)
            logger.debug(msg)
        i = 1
        affected_row_count = 0
        total_rows = len(prepared_list[1:])
        for row in prepared_list[1:]:
            query = f"INSERT INTO {table_name} ("
            values_query = f" VALUES ("
            for key, value in row.items():
                query += f" {key},"
                values_query += f" '{value}',"
            query = query[:-1]
            query += ")"
            values_query = values_query[:-1]
            values_query += ")"

            query = query+values_query+";"
            try:
                self.cursor.execute(query)
                self.connection.commit()
            except Error as e:
                row_msg = f"ROW = {i}"
                print(e)
                print(query)
                print(row_msg)
                logger.error(e)
                logger.debug(query)
                logger.debug(row_msg)
            msg = f"ROW AFFECTED = {self.cursor.rowcount}"
            # print(msg)
            # logger.debug(msg)
            i += 1
            affected_row_count += self.cursor.rowcount

        logger.debug(f"Total rows to be affected: {total_rows}")
        logger.debug(f"Total rows inserted/affected: {affected_row_count}")
        if(affected_row_count == (i-1)):
            logger.info(f"{table_name} updated successfully")
        else:
            logger.warning(f"{table_name} not updated successfully")

    def get_columns(self, table_name):
        self.cursor.execute(f"SELECT * FROM {table_name} LIMIT 0")
        desc_list = (self.cursor.description)
        columns = []
        for c in desc_list:
            columns.append(c[0])
        return set(columns)
